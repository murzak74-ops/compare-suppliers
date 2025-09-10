import streamlit as st
import pandas as pd

# --------------------------
# 1. Авторизация по e-mail
# --------------------------
AUTHORIZED_EMAILS = [
    "rab.org@bk.ru",
    "rab-organ@yandex.ru",
    "ooo.rab.org@gmail.com",
]

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.title("🔒 Доступ только для своих")
    email = st.text_input("Логин помнишь, да?")
    if st.button("Войти"):
        if email.strip().lower() in [e.lower() for e in AUTHORIZED_EMAILS]:
            st.session_state.authenticated = True
            st.success("Добро пожаловать ✅")
        else:
            st.error("❌ У вас нет доступа к этому приложению")
    st.stop()

# ---- Стилизация ----
st.markdown(
    """
    <style>
    .main { background-color: #f8f9fa; }
    h1, h2, h3 { font-family: 'Arial', sans-serif; }
    .instruction-box {
        background-color: #ffffff; border: 1px solid #dcdcdc; border-radius: 10px;
        padding: 15px 20px; margin-bottom: 20px; box-shadow: 2px 2px 6px rgba(0,0,0,0.05);
    }
    .stDataFrame tbody tr td { text-align: center; }
    .stDownloadButton button {
        background-color: #004080; color: white; border-radius: 8px; padding: 8px 16px; border: none;
    }
    .stDownloadButton button:hover { background-color: #0066cc; color: white; }
    </style>
    """,
    unsafe_allow_html=True,
)

# ---- Шапка с логотипом ----
col1, col2 = st.columns([1, 3])
with col1:
    st.image("Логонов2.png", width=200)
with col2:
    st.markdown(
        """
        <h1 style='color:#004080; font-size:38px; margin-bottom:0px;'>Сравнение цен поставщиков</h1>
        <h3 style='color:gray; font-weight:normal; margin-top:5px;'>Компания «Рабочий Орган»</h3>
        """,
        unsafe_allow_html=True,
    )

# ---- Инструкция ----
st.markdown(
    """
    <div class=\"instruction-box\">
    <b>Загрузите Excel-файл с колонками:</b><br>
    - <b>Артикул</b><br>
    - <b>Кол-во</b> (или <b>Количество</b>)<br>
    - пары колонок для каждого поставщика: <code>Цена_*</code> и <code>Производитель_*</code><br><br>
    <i>Пустые и нулевые цены не учитываются.</i>
    </div>
    """,
    unsafe_allow_html=True,
)

# ---- Образец заявки (скачать) ----
with st.container():
    st.markdown("**📎 Образец заявки** — скачайте и используйте как шаблон для отправки поставщикам.")

    @st.cache_data
    def build_request_template() -> bytes:
        # Отдаём загруженный пользователем файл-образец напрямую
        with open("Заявка_образец.xlsx", "rb") as f:
            return f.read()

    st.download_button(
        label="📥 Скачать образец заявки (Excel)",
        data=build_request_template(),
        file_name="Заявка_образец.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ---- Функционал ----
uploaded_file = st.file_uploader("Загрузите Excel", type=["xlsx", "xls"])


def parse_suppliers_columns(columns):
    """Находит пары колонок вида 'Цена_<Имя>' и 'Производитель_<Имя>'."""
    suppliers = {}
    for col in columns:
        if isinstance(col, str) and col.startswith("Цена_"):
            name = col.replace("Цена_", "")
            prod_col = f"Производитель_{name}"
            if prod_col in columns:
                suppliers[name] = (col, prod_col)
    return suppliers


if uploaded_file:
    try:
        df = pd.read_excel(uploaded_file)

        # Поддержка альтернативного имени колонки количества
        if "Кол-во" not in df.columns and "Количество" in df.columns:
            df = df.rename(columns={"Количество": "Кол-во"})
        if not {"Артикул", "Кол-во"}.issubset(df.columns):
            st.error("Файл должен содержать колонки 'Артикул' и 'Кол-во' (или 'Количество').")
            st.stop()

        suppliers = parse_suppliers_columns(df.columns)
        if not suppliers:
            st.error("Не найдены пары колонок вида 'Цена_*' и 'Производитель_*'.")
            st.stop()

        # В «длинный» формат (берём только валидные цены > 0)
        records = []
        for _, row in df.iterrows():
            for supplier, (price_col, prod_col) in suppliers.items():
                price = row[price_col]
                producer = row[prod_col]
                p = None
                if pd.notna(price):
                    try:
                        p = float(price)
                    except Exception:
                        p = None
                if p is not None and p > 0:
                    records.append({
                        "Артикул": row["Артикул"],
                        "Кол-во": row["Кол-во"],
                        "Поставщик": supplier,
                        "Цена": p,
                        "Производитель": producer,
                    })
        long_df = pd.DataFrame(records)

        # Режимы отображения
        mode = st.radio(
            "Выберите режим отображения:",
            ["Лучший поставщик", "Все поставщики (по возрастанию)"]
        )

        # Опция порядка групп
        group_by_original = st.checkbox(
            "Сначала аналоги, потом оригинал",
            value=False,
            help=(
                "В режиме 'Все поставщики' аналоги будут идти первыми, затем оригиналы "
                "(в каждой группе — по возрастанию цены)."
            ),
        )

        # Список базовых артикулов (сохраняем строки без цен)
        articles = df[["Артикул", "Кол-во"]].copy()

        if mode == "Лучший поставщик":
            rows = []
            for _, base in articles.iterrows():
                a = base["Артикул"]
                q = base["Кол-во"]
                g = long_df[long_df["Артикул"] == a]
                if not g.empty:
                    r = g.loc[g["Цена"].idxmin()]
                    rows.append({
                        "Артикул": a,
                        "Кол-во": q,
                        "Производитель": r["Производитель"],
                        "Поставщик": r["Поставщик"],
                        "Цена": r["Цена"],
                    })
                else:
                    rows.append({
                        "Артикул": a,
                        "Кол-во": q,
                        "Производитель": None,
                        "Поставщик": None,
                        "Цена": None,
                    })
            result = pd.DataFrame(rows)
            st.subheader("Лучшие цены по каждому артикулу")
            st.dataframe(result, use_container_width=True)
        else:
            # Все поставщики по возрастанию, в «широкую» строку
            def all_sorted_wide(group: pd.DataFrame) -> pd.Series:
                temp = group.copy()
                temp["__is_original"] = (
                    temp["Производитель"].astype(str).str.strip().str.lower().eq("оригинал")
                )
                if group_by_original:
                    temp = temp.sort_values(
                        ["__is_original", "Цена", "Поставщик"], ascending=[True, True, True]
                    )
                else:
                    temp = temp.sort_values(["Цена", "Поставщик"])
                row = {
                    "Артикул": group["Артикул"].iloc[0],
                    "Кол-во": group["Кол-во"].iloc[0],
                }
                for i, (_, r) in enumerate(temp.iterrows(), 1):
                    row[f"Цена_{i}"] = r["Цена"]
                    row[f"Поставщик_{i}"] = r["Поставщик"]
                    row[f"Производитель_{i}"] = r["Производитель"]
                return pd.Series(row)

            wide_rows = []
            for _, base in articles.iterrows():
                a = base["Артикул"]
                q = base["Кол-во"]
                g = long_df[long_df["Артикул"] == a]
                if not g.empty:
                    s = all_sorted_wide(g)
                    s["Артикул"], s["Кол-во"] = a, q
                    wide_rows.append(s.to_dict())
                else:
                    wide_rows.append({"Артикул": a, "Кол-во": q})
            result = pd.DataFrame(wide_rows).fillna(pd.NA)
            st.subheader("Все поставщики (по возрастанию цены)")
            st.dataframe(result, use_container_width=True)

        # Экспорт в Excel с форматированием
        @st.cache_data
        def to_excel_bytes(df_out: pd.DataFrame) -> bytes:
            from io import BytesIO
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter

            out = BytesIO()
            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                df_out.to_excel(writer, index=False, sheet_name="Результаты")
                ws = writer.sheets["Результаты"]

                # Закрепляем верхнюю строку и первый столбец
                ws.freeze_panes = "B2"
                # Включаем автофильтры
                ws.auto_filter.ref = ws.dimensions

                # Автоширина колонок
                for col_idx in range(1, ws.max_column + 1):
                    max_length = 0
                    col_letter = get_column_letter(col_idx)
                    for row_idx in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        value = cell.value
                        length = len(str(value)) if value is not None else 0
                        if length > max_length:
                            max_length = length
                    ws.column_dimensions[col_letter].width = max_length + 2

                # Выделение "Оригинал" жирным в колонках Производитель/Производитель_*
                header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                prod_cols = [
                    i + 1
                    for i, h in enumerate(header)
                    if isinstance(h, str) and (h == "Производитель" or h.startswith("Производитель_"))
                ]
                bold = Font(bold=True)
                for col_idx in prod_cols:
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if isinstance(cell.value, str) and cell.value.strip().lower() == "оригинал":
                            cell.font = bold

                # Числовой формат для колонок Цена/Цена_*
                price_cols = [
                    i + 1
                    for i, h in enumerate(header)
                    if isinstance(h, str) and (h == "Цена" or h.startswith("Цена_"))
                ]
                for col_idx in price_cols:
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "#,##0.00"

            out.seek(0)
            return out.getvalue()

        st.download_button(
            label="Скачать результат в Excel",
            data=to_excel_bytes(result),
            file_name=(
                "best_suppliers.xlsx" if mode == "Лучший поставщик" else "all_suppliers_sorted.xlsx"
            ),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(f"Ошибка при обработке файла: {e}")

