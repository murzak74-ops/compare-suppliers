import streamlit as st
import pandas as pd
import io
import os
import re
from typing import List, Optional, Dict, Tuple
# --------------------------
# 1. Авторизация по e-mail
# --------------------------
AUTHORIZED_EMAILS = [
    "rab.org@bk.ru",
    "rab-organ@yandex.ru",
    "ooo.rab.org@gmail.com",
]

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.title("🔒 Доступ только для своих")
    email = st.text_input("Введите ваш e-mail")
    if st.button("Войти"):
        if email.strip().lower() in [e.lower() for e in AUTHORIZED_EMAILS]:
            st.session_state.authenticated = True
            st.success("Добро пожаловать ✅")
        else:
            st.error("❌ У вас нет доступа к этому приложению")
    st.stop()

# ---- Стилизация ----

# ==============================
# VPR Importer (Standalone, v2)
# ==============================
# Задача: загрузить «базовую расценку» (список артикулов) и подтянуть к ней
# предложения из множества прайсов (Excel + цифровые PDF), выдать ОДНУ СТРОКУ на артикул
# с блоками Цена_i / Поставщик_i / Производитель_i по возрастанию цены.

st.set_page_config(page_title="VPR Importer", page_icon="🧩", layout="wide")

# ---------- PDF Support ----------
try:
    import pdfplumber  # type: ignore
    HAS_PDFPLUMBER = True
except Exception:
    HAS_PDFPLUMBER = False

# ---------- Styles ----------
st.markdown(
    """
    <style>
    .stDownloadButton button { background-color:#0d6efd; color:white; border-radius:8px; }
    .stDownloadButton button:hover { background-color:#0b5ed7; }
    .ok { color:#12b886; }
    .warn { color:#f08c00; }
    .err { color:#e03131; }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("🧩 VPR Importer — привязка к базовой расценке")
st.caption("Загрузите базовую расценку с артикулами, затем прайсы (Excel/PDF). Получите одну строку на артикул с вариантами цен.")

# ---------- Helpers ----------
PRICE_RE = re.compile(r"[\d\s.,]+")

COL_ART = "Артикул"
COL_QTY = "Кол-во"
COL_PRICE = "Цена"
COL_BRAND = "Производитель"
COL_VENDOR = "Поставщик"
COL_SRC = "Источник"
COL_NORM = "__ART_NORM"

SUPPORTED_HINTS = {
    COL_ART: ["артикул", "код", "sku", "part", "номер"],
    COL_PRICE: ["цена", "price", "стоим", "cost"],
    COL_BRAND: ["производ", "бренд", "brand", "maker"],
    COL_QTY: ["кол-во", "количество", "qty", "колич"],
}

def normalize_part(s: str) -> str:
    if not isinstance(s, str):
        s = str(s)
    return re.sub(r"[^A-Z0-9]", "", s.upper())


def parse_price(val, decimal:"," = ",") -> Optional[float]:
    # Сначала числа как есть
    if isinstance(val, (int, float)) and not pd.isna(val):
        f = float(val)
        return f if f > 0 else None
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val)
    m = PRICE_RE.search(s)
    if not m:
        return None
    num = m.group(0).replace(" ", "").replace("\xa0", "")
    if decimal == ",":
        parts = num.rsplit(",", 1)
        if len(parts) == 2:
            num = parts[0].replace(".", "") + "." + parts[1]
        else:
            num = num.replace(".", "")
    else:
        num = num.replace(",", "")
    try:
        f = float(num)
        return f if f > 0 else None
    except Exception:
        return None


def suggest_column(columns: List[str], hints: List[str]) -> Optional[str]:
    lc_map = {str(c).strip().lower(): c for c in columns}
    for hint in hints:
        for lc, orig in lc_map.items():
            if hint in lc:
                return orig
    return None


def parse_excel(file_bytes: bytes) -> pd.DataFrame:
    return pd.read_excel(io.BytesIO(file_bytes))


def parse_pdf_tables(file_bytes: bytes) -> List[pd.DataFrame]:
    if not HAS_PDFPLUMBER:
        return []
    frames = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for tbl in tables:
                if not tbl or len(tbl) < 2:
                    continue
                headers = [str(h).strip() if h is not None else "" for h in tbl[0]]
                rows = tbl[1:]
                df = pd.DataFrame(rows, columns=headers)
                frames.append(df)
    return frames


def normalize_rows(df: pd.DataFrame, art_col: str, price_col: str, brand_col: Optional[str], vendor: str, src_label: str, decimal_sep: str) -> pd.DataFrame:
    out_rows = []
    for _, r in df.iterrows():
        part = r.get(art_col)
        price = parse_price(r.get(price_col), decimal_sep)
        if price is None:
            continue
        brand = r.get(brand_col) if (brand_col and brand_col in df.columns) else None
        out_rows.append({
            COL_ART: part,
            COL_PRICE: price,
            COL_BRAND: brand,
            COL_VENDOR: vendor,
            COL_SRC: src_label,
        })
    out = pd.DataFrame(out_rows)
    if not out.empty:
        out = out.dropna(subset=[COL_ART]).copy()
        out[COL_ART] = out[COL_ART].astype(str)
        out[COL_NORM] = out[COL_ART].apply(normalize_part)
    return out

# =============================
# 1) БАЗОВАЯ РАСЦЕНКА (обязательно)
# =============================
st.subheader("1) Загрузите базовую расценку")
st.caption("Минимум одна колонка с артикулами. Колонка количества — опционально.")
base_file = st.file_uploader("Базовая расценка (Excel)", type=["xlsx", "xls"], key="base")
base_df = None
if base_file:
    try:
        base_raw = parse_excel(base_file.read())
        cols = list(base_raw.columns)
        art_col = suggest_column(cols, SUPPORTED_HINTS[COL_ART]) or cols[0]
        qty_col = suggest_column(cols, SUPPORTED_HINTS[COL_QTY])
        c1, c2 = st.columns(2)
        with c1:
            art_col = st.selectbox("Колонка артикула", options=cols, index=(cols.index(art_col) if art_col in cols else 0))
        with c2:
            qty_col = st.selectbox("Колонка количества (опционально)", options=["<нет>"] + cols, index=(0 if qty_col is None else cols.index(qty_col)+1))
        base_df = pd.DataFrame({COL_ART: base_raw[art_col].astype(str)})
        if qty_col != "<нет>":
            base_df[COL_QTY] = base_raw[qty_col]
        else:
            base_df[COL_QTY] = None
        base_df[COL_NORM] = base_df[COL_ART].apply(normalize_part)
        st.success(f"Загружено позиций: {len(base_df)}")
        st.dataframe(base_df.head(30), use_container_width=True)
    except Exception as e:
        st.error(f"Ошибка чтения базовой расценки: {e}")

# =============================
# 2) ПРАЙСЫ ПОСТАВЩИКОВ (VPR)
# =============================
st.subheader("2) Загрузите прайсы поставщиков (VPR)")
st.caption("Поддерживаются Excel (XLS/XLSX) и цифровые PDF. Сканам требуется OCR (не входит).")
vpr_files = st.file_uploader("Прайсы (Excel/PDF)", type=["xlsx","xls","pdf"], accept_multiple_files=True, key="vprs")
decimal_sep = st.selectbox("Десятичный разделитель в ценах VPR", [",", "."], index=0)
try_pdf = st.checkbox("Извлекать таблицы из PDF", value=True and HAS_PDFPLUMBER)

all_offers: List[pd.DataFrame] = []
if vpr_files:
    for f in vpr_files:
        with st.expander(f"📄 {f.name}", expanded=True):
            vendor_default = os.path.splitext(f.name)[0]
            vendor_val = st.text_input("Имя поставщика", value=vendor_default, key=f"vendor::{f.name}")
            file_bytes = f.read()
            src_label = f.name

            if f.name.lower().endswith((".xlsx",".xls")):
                try:
                    df = parse_excel(file_bytes)
                except Exception as e:
                    st.error(f"Ошибка чтения Excel: {e}")
                    continue
                cols = list(df.columns)
                art_guess = suggest_column(cols, SUPPORTED_HINTS[COL_ART]) or (cols[0] if cols else None)
                price_guess = suggest_column(cols, SUPPORTED_HINTS[COL_PRICE]) or (cols[1] if len(cols)>1 else None)
                brand_guess = suggest_column(cols, SUPPORTED_HINTS[COL_BRAND])
                c1,c2,c3 = st.columns(3)
                with c1:
                    art_col = st.selectbox("Столбец артикула", options=cols, index=(cols.index(art_guess) if art_guess in cols else 0), key=f"art::{f.name}")
                with c2:
                    price_col = st.selectbox("Столбец цены", options=cols, index=(cols.index(price_guess) if price_guess in cols else (1 if len(cols)>1 else 0)), key=f"price::{f.name}")
                with c3:
                    brand_col = st.selectbox("Столбец производителя", options=["<нет>"]+cols, index=(0 if brand_guess is None else cols.index(brand_guess)+1), key=f"brand::{f.name}")
                offers = normalize_rows(df, art_col, price_col, (None if brand_col=="<нет>" else brand_col), vendor_val, src_label, decimal_sep)
                st.write(f"Найдено строк с ценой: **{len(offers)}**")
                if not offers.empty:
                    st.dataframe(offers.head(20), use_container_width=True)
                    all_offers.append(offers)
                else:
                    st.warning("Не удалось распознать цены. Проверьте выбор колонок и десятичный разделитель.")

            elif f.name.lower().endswith(".pdf"):
                if not (try_pdf and HAS_PDFPLUMBER):
                    st.warning("PDF не обработан: нет pdfplumber или выключено извлечение.")
                else:
                    try:
                        tables = parse_pdf_tables(file_bytes)
                    except Exception as e:
                        st.error(f"Ошибка чтения PDF: {e}")
                        tables = []
                    if not tables:
                        st.warning("Таблицы в PDF не найдены.")
                    for idx, df in enumerate(tables, start=1):
                        with st.expander(f"Таблица {idx}"):
                            cols = list(df.columns)
                            if not cols:
                                st.warning("Пустая таблица.")
                                continue
                            art_guess = suggest_column(cols, SUPPORTED_HINTS[COL_ART]) or cols[0]
                            price_guess = suggest_column(cols, SUPPORTED_HINTS[COL_PRICE]) or (cols[1] if len(cols)>1 else cols[0])
                            brand_guess = suggest_column(cols, SUPPORTED_HINTS[COL_BRAND])
                            c1,c2,c3 = st.columns(3)
                            with c1:
                                art_col = st.selectbox("Столбец артикула", options=cols, index=(cols.index(art_guess) if art_guess in cols else 0), key=f"pdf_art::{f.name}::{idx}")
                            with c2:
                                price_col = st.selectbox("Столбец цены", options=cols, index=(cols.index(price_guess) if price_guess in cols else (1 if len(cols)>1 else 0)), key=f"pdf_price::{f.name}::{idx}")
                            with c3:
                                brand_col = st.selectbox("Столбец производителя", options=["<нет>"]+cols, index=(0 if brand_guess is None else cols.index(brand_guess)+1), key=f"pdf_brand::{f.name}::{idx}")
                            offers = normalize_rows(df, art_col, price_col, (None if brand_col=="<нет>" else brand_col), vendor_val, f"{src_label} :: Таблица {idx}", decimal_sep)
                            st.write(f"Найдено строк с ценой: **{len(offers)}**")
                            if not offers.empty:
                                st.dataframe(offers.head(20), use_container_width=True)
                                all_offers.append(offers)
                            else:
                                st.warning("В этой таблице цены не распознаны.")

# =============================
# 3) СВЯЗЫВАНИЕ С БАЗОЙ -> WIDE
# =============================
if base_df is None:
    st.info("Загрузите базовую расценку (п.1).")
    st.stop()

if not all_offers:
    st.info("Загрузите хотя бы один прайс (п.2).")
    st.stop()

offers_df = pd.concat(all_offers, ignore_index=True).dropna(subset=[COL_ART])
offers_df[COL_ART] = offers_df[COL_ART].astype(str)
offers_df[COL_NORM] = offers_df[COL_ART].apply(normalize_part)

# join по нормализованному артикулу — оставляем только то, что есть в базе
base_norm = base_df[[COL_ART, COL_QTY, COL_NORM]].drop_duplicates()
matched = offers_df.merge(base_norm[[COL_NORM, COL_QTY, COL_ART]], on=COL_NORM, how="inner", suffixes=("_offer","_base"))
if matched.empty:
    st.warning("Совпадений по артикулам не найдено. Проверьте формат артикула в базе и прайсах.")

# wide-преобразование: одна строка на артикул базы
st.markdown("---")
st.subheader("Итог: одна строка на артикул (цены по возрастанию)")

def build_wide_full(base_df: pd.DataFrame, matched: pd.DataFrame) -> pd.DataFrame:
    # сгруппуем найденные предложения по нормализованному артикулу
    offers_by_norm: Dict[str, pd.DataFrame] = {}
    if not matched.empty:
        for norm, g in matched.groupby(COL_NORM):
            offers_by_norm[norm] = g.sort_values(COL_PRICE)

    rows = []
    max_slots = 0  # сколько блоков цена/поставщик/производитель понадобится максимум

    for _, b in base_df.iterrows():
        norm = b[COL_NORM]
        row = {COL_ART: b[COL_ART], COL_QTY: b[COL_QTY]}
        g = offers_by_norm.get(norm)

        if g is not None and not g.empty:
            for i, (_, r) in enumerate(g.iterrows(), 1):
                row[f"{COL_PRICE}_{i}"]  = r[COL_PRICE]
                row[f"{COL_VENDOR}_{i}"] = r[COL_VENDOR]
                row[f"{COL_BRAND}_{i}"]  = r[COL_BRAND]
            max_slots = max(max_slots, len(g))
        # если предложений нет — оставляем только Артикул/Кол-во (цены пустые)
        rows.append(row)

    wide = pd.DataFrame(rows)

    # выровняем и упорядочим колонки: Артикул, Кол-во, [Цена_i, Поставщик_i, Производитель_i]...
    # добавим недостающие колонки с пустыми значениями
    for i in range(1, max_slots + 1):
        for col in (f"{COL_PRICE}_{i}", f"{COL_VENDOR}_{i}", f"{COL_BRAND}_{i}"):
            if col not in wide.columns:
                wide[col] = pd.NA

    ordered = [COL_ART, COL_QTY] + sum(
        ([f"{COL_PRICE}_{i}", f"{COL_VENDOR}_{i}", f"{COL_BRAND}_{i}"] for i in range(1, max_slots + 1)),
        []
    )
    return wide.reindex(columns=ordered)

wide = build_wide_full(base_df, matched)
st.dataframe(wide, use_container_width=True)

# ==================
# 4) Экспорт в Excel
# ==================
from io import BytesIO
@st.cache_data
def df_to_xlsx_bytes(df_out: pd.DataFrame) -> bytes:
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="VPR")
        ws = writer.sheets["VPR"]
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        # автоширина
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                max_len = max(max_len, len(str(v)) if v is not None else 0)
            ws.column_dimensions[col_letter].width = max_len + 2
        # формат цен
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        price_cols = [i+1 for i,h in enumerate(header) if isinstance(h,str) and (h==COL_PRICE or h.startswith(f"{COL_PRICE}_"))]
        for c in price_cols:
            for r in range(2, ws.max_row+1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, (int,float)):
                    cell.number_format = "#,##0.00"
        # жирный "Оригинал"
        brand_cols = [i+1 for i,h in enumerate(header) if isinstance(h,str) and (h==COL_BRAND or h.startswith(f"{COL_BRAND}_"))]
        bold = Font(bold=True)
        for c in brand_cols:
            for r in range(2, ws.max_row+1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, str) and cell.value.strip().lower()=="оригинал":
                    cell.font = bold
    bio.seek(0)
    return bio.getvalue()

st.download_button(
    label="📥 Скачать результат (Excel)",
    data=df_to_xlsx_bytes(wide),
    file_name="vpr_wide_by_base.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
